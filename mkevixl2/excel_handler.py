from openpyxl import Workbook
import os

"""Excelに対する処理を行う"""


def make_new_workbook_name(filename):
    """Excelファイル名を返す

    Args:
        filename ([type]): [description]
    """

    if not filename.endswith(".xlsx"):
        # 拡張子がxlsx以外だったら強制的に変更する
        filename = os.path.splitext(filename)[0] + ".xlsx"

    return filename


class ExcelSheetByName:
    """指定した名称でシートを追加してExcelファイル保存"""

    def __init__(self, filename, sheetnames):
        """[summary]

        Args:
        filename ([string]): [excelfilename]
        sheetnames ([list]): [sheetname]
        """
        self.filename = filename
        self.sheetnames = sheetnames

    def execute(self):

        wb = Workbook()
        for sheetname in self.sheetnames:
            wb.create_sheet(sheetname)
        wb.remove(wb["Sheet"])

        return wb.save(self.filename)


class ExcelSheetByCount:
    """[指定した数字名と数だけシートを追加してExcelファイル保存]"""

    def __init__(self, filename, start, sheetcount):
        """[summary]

        Args:
            filename ([string]): [ファイル名]
            start ([int]): [開始値]
            sheetcount ([int]): [追加シート数]
        """
        self.filename = filename
        self.start = start
        self.sheetcount = sheetcount

    def execute(self):

        sheetname = self.start
        # print(sheetname + 1)
        wb = Workbook()
        for count in range(0, int(self.sheetcount)):
            wb.create_sheet(str(sheetname))
            sheetname += 1
        wb.remove(wb["Sheet"])

        return wb.save(self.filename)


def make_file(filename, sheetnames):

    return ExcelSheetByName(make_new_workbook_name(filename), sheetnames).execute()
